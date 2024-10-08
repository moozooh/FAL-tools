import datetime
import aiohttp
from aiohttp import ClientResponseError, ClientError
import asyncio
from asyncio import Semaphore
from contextlib import asynccontextmanager
import sys
import math
import re
import logging
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Alignment, Font, numbers
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from bs4 import BeautifulSoup

# v v v v v SET USER VARIABLES HERE v v v v v

# MAL & FAL SETTINGS
# Insert the client ID you received from MAL in the quotes
client_id = ''

# Set a short name to identify target FAL season in output filenames
season_code = '24fall'

# Set ace eligibility threshold (the "cap") for the season
ace_cap = 70000

# Set to a 'YYYY-MM-DD' string in quotation marks to override season start date. The day should be Monday of the first week in the UTC timezone; the script will convert it to FAL timezone automatically.
season_start_override = None  # Default: None

# Set season name in quotation marks if you override the date.
season_override = None  # Default: None

# DATA COLLECTION SETTINGS
# Populate this list with IDs of the anime to track. The ID can be found in the URL for the anime page on MAL, e.g. for Trigun (https://myanimelist.net/anime/6/Trigun) the ID is "6". The following list is only relevant for fall 24
anime_ids = [50306, 52215, 52995, 53033, 53287, 54726, 54853, 55071, 55150, 55823, 55887, 55994, 56228, 56647, 56784, 56843, 56894, 56964, 57066, 57181, 57360, 57533, 57554, 57611, 57635, 57891, 57944, 58172, 58572, 58714, 59131, 59145, 59425]

# Optional secondary list of anime that need tracking. 
secondary_anime_ids = []  # Default: []

# Enable forum post data fetching. This compiles all forum post statistics, including the names and dates of unique user posts, on a separate sheet. WARNING: The code enabled by this function takes a full snapshot of the forum data for each anime ID specified. With many shows, especially late in the season, this hammers the site with hundreds of API and HTTP requests at once. Server-side throttling is inevitable, and the output filesize and network traffic will grow by orders of magnitude. Avoid using more than once per hour.
enable_posts = False  # Default: False

# SHEET SETTINGS
# Set the column to sort the sheet by. Columns 'A' (Title) and 'M' (ID) are recommended for consistency when passing/manipulating the data further
sort_column = 'M'  # Default: 'M'

# Set to True to hide the "Watching" column
hide_watching = True  # Default: True

# Set to True to hide the "Completed" column
hide_completed = True  # Default: True

# Set to True to hide the "ID" column
hide_id = True  # Default: True

# CONSOLE SETTINGS
# Adjust console output verbosity level.
# 0: silent (no output, no logging);
# 1: basic (progress report and errors);
# 2: verbose (detailed action report);
# 3: extremely verbose (for debugging purposes only).
verbosity = 1  # Default: 1

# Enable logging to a text file. Use to keep track of scheduled operation. The log is created in the same folder as the script and has the same name with a .log extension. Requires verbosity levels 1, 2, or 3; won't do anything at 0
enable_logging = False  # Default: False

# NETWORK SETTINGS
# Limit concurrent asynchronous requests to avoid server-side throttling. Will not do much if your usage snapshots data in intervals of 6 minutes or greater. Will also not do much if you're fetching data for more than 10 titles at a time or have post data fetching enabled. Default value is practically unlimited
max_conc_requests = 9999  # Default: 9999

# Number of attempts to retry a failed API or HTTP call
retry_limit = 3  # Default: 3

# Time to wait between each retry in seconds
sleep_time = 20  # Default: 20

# ^ ^ ^ ^ ^ ^ END USER VARIABLES ^ ^ ^ ^ ^ ^

# Add a custom logging level for verbosity level 2 between INFO and DEBUG
VERBOSE = 15
logging.addLevelName(VERBOSE, "VERBOSE")

# Set up logging based on verbosity level and logging preferences
if enable_logging and verbosity > 0:
    log_filename = __file__.rsplit('.', 1)[0] + '.log'
    log_level = logging.INFO if verbosity == 1 else VERBOSE if verbosity == 2 else logging.DEBUG
    logging.basicConfig(
        filename=log_filename,
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    logger = logging.getLogger()
else:
    logger = None

# Initialize an error counter and an array for IDs we failed to fetch
error_count = 0
failed_anime_ids = []

# Set up the logic for verbosity levels
def log_print(message, level=logging.INFO):
    global error_count
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    if level == logging.ERROR:
        error_count += 1

    if verbosity == 0:
        return
    elif verbosity == 1 and level >= logging.INFO:
        print(f"{timestamp}: {message}")
    elif verbosity == 2 and level >= VERBOSE:
        print(f"{timestamp}: {message}")
    elif verbosity >= 3:
        print(f"{timestamp}: {logging.getLevelName(level)} - {message}")
    
    if logger:
        logger.log(level, message)

# Print verbosity indicators for levels 2 and 3
log_print("Selected verbosity level will print out full action report", VERBOSE)  # Printed at verbosity levels 2 and 3
log_print("Selected verbosity level will also print out full debug information, including the content of almost every array, variable, and API response", logging.DEBUG)  # Printed only at verbosity level 3

# Start documenting script operation
log_print(f"Script started", logging.INFO)
if enable_logging and verbosity > 0:
    log_print(f"Logging output to: {log_filename}", logging.INFO)

# Print the status of the post fetching variable
if enable_posts:
    log_print("Post data fetching and calculation enabled", logging.INFO)
else:
    log_print("Post data fetching and calculation disabled", logging.INFO)

# Validate that all provided anime IDs are numeric and the list is not empty
def validate_anime_ids(ids):
    return all(str(id).isdigit() for id in ids) and len(ids) > 0

# Display error if IDs not found or invalid
if not validate_anime_ids(anime_ids):
    log_print("Error: Anime IDs missing or invalid. Check the user variables section of the script.", level=logging.ERROR)
    sys.exit(1)

# Set up asynchronous data fetching with client-side throttling
async def fetch_anime_data(session, anime_id, semaphore):
    async with semaphore:
        url = f'https://api.myanimelist.net/v2/anime/{anime_id}?fields=mean,num_favorites,statistics,status'
        headers = {'X-MAL-CLIENT-ID': client_id}
        attempts = max(1, retry_limit)  # Ensure at least one attempt is made
        for attempt in range(attempts):
            try:
                async with session.get(url, headers=headers) as response:
                    response.raise_for_status()
                    return await response.json()
            except ClientResponseError as e:
                if e.status == 404:
                    log_print(f"Anime with ID {anime_id} not found (error 404)", level=logging.ERROR)
                    return None 
                elif e.status == 504 or attempt == attempts - 1:
                    log_print(f"Failed to fetch anime data for ID {anime_id} after {attempt + 1} attempts: {e.status}, message='{e.message}', url='{e.request_info.url}'", level=logging.ERROR)
                    return None
                await asyncio.sleep(sleep_time)
            except ClientError as e:
                if attempt == attempts - 1:
                    log_print(f"Failed to fetch anime data for ID {anime_id} after {retry_limit} attempts: {str(e)}", level=logging.ERROR)
                    return None
                if retry_limit > 0:
                    await asyncio.sleep(sleep_time)
            except Exception as e:
                if attempt == attempts - 1:
                    log_print(f"Unexpected error fetching data for anime ID {anime_id}: {str(e)}", level=logging.ERROR)
                    return None
                if retry_limit > 0:
                    await asyncio.sleep(sleep_time)
    return None  # If all attempts fail

# Ensure optimal usage of client sessions by using a context manager
@asynccontextmanager
async def get_session():
    async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=False)) as session:
        yield session

# Main asynchronous function to fetch data for all provided anime IDs
async def main(anime_ids):
    async with get_session() as session:
        semaphore = Semaphore(max_conc_requests)
        tasks = [fetch_anime_data(session, anime_id, semaphore) for anime_id in anime_ids]
        responses = await asyncio.gather(*tasks)
        # Create a dictionary mapping anime IDs to their data (or None if fetch failed)
        return {anime_id: response for anime_id, response in zip(anime_ids, responses) if response is not None}

# Set up the Windows event loop policy to avoid errors if running on Windows
if sys.platform.startswith('win'):
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

# Run the asynchronous fetching for primary list, document the action
log_print("Fetching API data...", logging.INFO)
datas = asyncio.run(main(anime_ids))
log_print(datas, logging.DEBUG)
failed_anime_ids.extend([anime_id for anime_id in anime_ids if anime_id not in datas])

# Run the asynchronous fetching for secondary list, if provided
if secondary_anime_ids:
    secondary_datas = asyncio.run(main(secondary_anime_ids))
    log_print(secondary_datas, logging.DEBUG)
else:
    secondary_datas = []

# Set up airing status codes and their respective fill colors
def set_status_code_and_fill(cell, status_code):
    status_colors = {
        'NYA': 'E0E0E0',  # Light grey
        'AIR': 'D8F2F2',  # Light blue
        'PRE': 'FFD7AB',  # Sand yellow
        'FIN': 'ABF1AB',  # Mint green
        'FIN?': 'FF4D4D'  # Bright red
    }
    cell.value = status_code
    cell.alignment = Alignment(horizontal='center')
    if status_code in status_colors:
        cell.fill = PatternFill(start_color=status_colors[status_code], end_color=status_colors[status_code], fill_type='solid')

# Create and set up the output Excel file
headers = ['Title', 'Score', 'Favorites', 'Posts', 'Watching', 'Completed', 'W+C', 'Dropped', 'Drop Rate', 'PTW', 'PTW Ratio', 'Status', 'ID']
workbook = Workbook()
main_sheet = workbook.active
main_sheet.title = 'main'

# Calculate and process useful variables for logic and math operations below
def calculate_variables(data):
    stats = data.get('statistics', {}).get('status', {})
    status = data.get('status', '')
    watching = int(stats.get('watching', 0))  # Default to 0 if null/missing
    completed = int(stats.get('completed', 0))  # \
    dropped = int(stats.get('dropped', 0))      #  > Likewise
    ptw = int(stats.get('plan_to_watch', 0))    # /

    # Set up heuristic for detecting early airings conflicting with MAL dates
    preair_noise_floor = math.ceil(200 + ptw // 250)

    # Set up heuristic for detecting early endings conflicting with MAL dates
    comp_noise_floor = math.ceil(100 + watching / 30)

    # Apply status detection heuristics and discard all noise
    if status == 'not_yet_aired':
        if watching < preair_noise_floor:
            watching, completed, dropped = 0, 0, 0
            status_code = 'NYA'
        else:
            completed = 0
            status_code = 'PRE'  # Mark as early preview
    elif status == 'finished_airing':
        status_code = 'FIN'
    elif status == 'currently_airing':
        if completed < comp_noise_floor:
            completed = 0
            status_code = 'AIR'
        else:
            status_code = 'FIN?'  # Mark as assumed finished
    else:
        status_code = 'UNK'  # Unknown status; shouldn't ever happen

    # Calculate and return processed values
    watch_comp = watching + completed
    watch_drop = watch_comp + dropped
    active_ratio = watch_comp / ptw if ptw > 0 else 0.00
    return status_code, watching, completed, dropped, ptw, watch_comp, watch_drop, active_ratio

# Sheet sorting logic
def sort_data(datas, sort_column):
    # Helper function to get a sortable value
    def get_sortable_value(x, key):
        if isinstance(x, str):  # Error message
            return float('inf')  # Ensure error messages are at the bottom
        return x.get(key, float('inf'))

    if sort_column == 'A':    # Title: alphabetic, A to Z
        return sorted(datas.values(), key=lambda x: x if isinstance(x, str) else x['title'].lower())
    elif sort_column == 'B':  # Score: higher to lower to null
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), -get_sortable_value(x, 'mean') if not isinstance(x, str) else 0))
    elif sort_column == 'C':  # Favorites: higher to lower
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), -get_sortable_value(x, 'num_favorites')))
    elif sort_column == 'D':  # Posts: higher to lower
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), -post_counts.get(int(x['id']), 0) if not isinstance(x, str) else 0))
    elif sort_column == 'E':  # Watching: higher to lower
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), -calculate_variables(x)[1] if not isinstance(x, str) else 0))
    elif sort_column == 'F':  # Completed: higher to lower
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), -calculate_variables(x)[2] if not isinstance(x, str) else 0))
    elif sort_column == 'G':  # W+C: higher to lower
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), -calculate_variables(x)[5] if not isinstance(x, str) else 0))
    elif sort_column == 'H':  # Dropped: higher to lower
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), -calculate_variables(x)[3] if not isinstance(x, str) else 0))
    elif sort_column == 'I':  # Drop Rate: higher to lower
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), -(calculate_variables(x)[3] / calculate_variables(x)[6]) if not isinstance(x, str) and calculate_variables(x)[6] > 0 else 0))
    elif sort_column == 'J':  # PTW: higher to lower
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), -calculate_variables(x)[4] if not isinstance(x, str) else 0))
    elif sort_column == 'K':  # PTW Ratio: higher to lower
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), -calculate_variables(x)[7] if not isinstance(x, str) else 0))
    elif sort_column == 'L':  # Status: FIN > FIN? > AIR > PRE > NYA > ERR
        status_order = {'FIN': 0, 'FIN?': 1, 'AIR': 2, 'PRE': 3, 'NYA': 4, 'ERR': 5}
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), status_order.get(calculate_variables(x)[0], 4) if not isinstance(x, str) else 5))
    elif sort_column == 'M':  # ID: lower to higher
        return sorted(datas.values(), key=lambda x: (isinstance(x, str), int(x['id']) if not isinstance(x, str) else float('inf')))

    # If no valid sort column is specified, return the original data
    return list(datas.values())

# Function to populate the main sheet with collected data
def populate_sheet(sheet, datas):
    # Format and freeze the headers and first column
    for i, header in enumerate(headers):
        cell = sheet.cell(1, i + 1, value=header)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
    sheet.freeze_panes = 'B2'

    # Populate data columns
    for j, data in enumerate(datas):
        if isinstance(data, str):
            # Write error message in all cells for this row
            for col in range(1, len(headers) + 1):
                cell = sheet.cell(j + 2, col, value=data)
                cell.font = Font(color="FF0000")  # Make text red
            continue

        sheet.cell(j + 2, 1, value=str(data['title']))
        sheet.cell(j + 2, 2, value=data.get("mean") or ' ').number_format = numbers.FORMAT_NUMBER_00
        sheet.cell(j + 2, 3, value=int(data['num_favorites']))
        # Initialize the 'Posts' column to be filled conditionally
        sheet.cell(j + 2, 4, value=0)
        # Extract processed values from calculate_variables()
        try:
            status_code, watching, completed, dropped, ptw, watch_comp, watch_drop, active_ratio = calculate_variables(data)
        except Exception as e:
            log_print(f"Error calculating variables for anime {data['title']}: {str(e)}", level=logging.ERROR)
            status_code, watching, completed, dropped, ptw, watch_comp, watch_drop, active_ratio = ("ERR", 0, 0, 0, 0, 0, 0, 0)
        sheet.cell(j + 2, 5, value=watching)
        sheet.cell(j + 2, 6, value=completed)
        sheet.cell(j + 2, 7, value=watch_comp)
        sheet.cell(j + 2, 8, value=dropped)
        # Avoid division-by-zero errors
        sheet.cell(j + 2, 9, value=dropped / watch_drop if watch_drop > 0 else 0).number_format = numbers.FORMAT_PERCENTAGE_00
        sheet.cell(j + 2, 10, value=ptw)
        sheet.cell(j + 2, 11, value=active_ratio).number_format = numbers.FORMAT_PERCENTAGE_00
        # Invoke the airing status heuristic to help determine the status
        status_cell = sheet.cell(j + 2, 12)
        set_status_code_and_fill(status_cell, status_code)
        sheet.cell(j + 2, 13, value=int(data['id'])).alignment = Alignment(horizontal='center')

    # Double the width of title column for better legibility by default
    sheet.column_dimensions['A'].width = 2 * sheet.column_dimensions['A'].width

# Sort the data before populating the sheet
sorted_datas = sort_data(datas, sort_column)

# Populate the main sheet
log_print("Populating the spreadsheet...", logging.INFO)
populate_sheet(main_sheet, sorted_datas)
# Populate the secondary sheet if secondary IDs are provided
if secondary_anime_ids:
    alt_sheet = workbook.create_sheet(title='alt')
    secondary_datas = sort_data(secondary_datas, sort_column)
    populate_sheet(alt_sheet, secondary_datas)

# Define conditional formatting and column hiding
def apply_conditional_formatting(sheet):
    green_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    red_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    # Column I fill goes from green (≤0.4%) to white (3.8%) to red (≥10%)
    color_rule_i = ColorScaleRule(start_type='num', start_value=0.004, start_color='548235', mid_type='num', mid_value=0.038, mid_color='FFFFFF', end_type='num', end_value=0.1, end_color='C00000')
    sheet.conditional_formatting.add('I2:I101', color_rule_i)
    sheet.conditional_formatting.add('I2:I101', CellIsRule(operator='greaterThanOrEqual', formula=['0.1'], fill=red_fill))

    # Color text in columns E, F, G mahogany if ace cap is reached
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=7):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value >= ace_cap:
                cell.font = Font(color="640D0D")

    # Skip empty cells and zero values
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=9, max_col=9):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value > 0:
                sheet.conditional_formatting.add(cell.coordinate, CellIsRule(operator='lessThanOrEqual', formula=['0.004'], fill=green_fill))

    # Column K goes from deep blue (lowest) to white (100%) to red (highest)
    color_rule_k = ColorScaleRule(start_type='min', start_color='1874A5', end_type='max', end_color='C00000', mid_type='num', mid_value=1.0, mid_color='FFFFFF')
    sheet.conditional_formatting.add('K2:K101', color_rule_k)

    # Add on-hover comments explaining columns and their formatting
    sheet['G1'].comment = Comment("Sum of watching and completed users. Ignores completed users if the series is not detected as completed, ignores both if the series is not detected as started.", " ")
    sheet['G1'].comment.width = None
    sheet['I1'].comment = Comment("Dropped users taken as a percentage of total dropped + watching + completed users.\nLower is better. Peak positive reached at 0.4% and below, peak negative at 10.0% and above, midpoint is at 3.8%.", " ")
    sheet['I1'].comment.width = None
    sheet['K1'].comment = Comment("Ratio of watching + completed users to PTW users.\nHigher (hot) values suggest good conversion from potential to active audience.\nMidpoint is at 100%.", " ")
    sheet['K1'].comment.width = None

    # Hide columns based on user variables
    if not enable_posts:
        sheet.column_dimensions['D'].hidden = True
    if hide_watching:
        sheet.column_dimensions['E'].hidden = True
    if hide_completed:
        sheet.column_dimensions['F'].hidden = True
    if hide_id:
        sheet.column_dimensions['M'].hidden = True
    # Always hide Posts on the alt sheet since we aren't populating it
    if sheet.title == 'alt':
        sheet.column_dimensions['D'].hidden = True

# Apply conditional formatting to relevant sheets
apply_conditional_formatting(main_sheet)
if secondary_datas:
    apply_conditional_formatting(alt_sheet)

# Set UTC+2 (FAL timezone) as the working timezone
now = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=2)

# Determine TV season boundaries based on the first week of the year
def get_season_start(year, season_start_override, season_override):
    if season_start_override and season_override:
         # Parse the override date and enforce UTC+2
        override_date = datetime.datetime.strptime(season_start_override, '%Y-%m-%d')
        return {season_override: override_date.replace(tzinfo=datetime.timezone.utc) + datetime.timedelta(hours=2)}
    else:
        # Week 1 to start on the Monday of the week containing 4 Jan
        iso_year_start = datetime.datetime(year, 1, 4, tzinfo=datetime.timezone.utc)
        iso_week_start = iso_year_start - datetime.timedelta(days=iso_year_start.weekday())
        # Divide the year into programming seasons by counting weeks
        season_week_starts = {
            'winter': 0,
            'spring': 13,
            'summer': 26,
            'fall': 39,
        }
        return {season: iso_week_start + datetime.timedelta(weeks=weeks) for season, weeks in season_week_starts.items()}

# Determine the current season and its start date based on current date
def get_current_season_and_start(now, season_start_override=None, season_override=None):
    year = now.year
    season_starts = get_season_start(year, season_start_override, season_override)
    
    for season, start_date in season_starts.items():
        end_date = start_date + datetime.timedelta(weeks=13)
        if start_date <= now < end_date:
            return season, start_date
    # If not found in the current year, assume to be next winter
    return 'winter', get_season_start(year + 1, season_start_override, season_override)['winter']

# Run the calculations
current_season, season_start = get_current_season_and_start(now, season_start_override, season_override)

# Calculate the current week and period within the season
def get_week_and_period(date, season_start):
    # Calculate the number of days since the start of the season
    delta_days = (date - season_start).days
    # Determine the current week
    current_week = (delta_days // 7) + 1
    # Determine the current period
    if current_week > 13 or current_week < 1:
        current_period = 'n/a'
    elif current_week == 13:
        current_period = 13
    else:
        current_period = min((current_week + 1) // 2 * 2, 12)  # Ensure period doesn't exceed 12
    
    return current_week, current_period

# Run the calculations
current_week, current_period = get_week_and_period(now, season_start)

# Calculate the timestamps for current period's start and end boundaries
if current_period == 2:
    current_period_start = season_start
else:
    current_period_start = season_start + datetime.timedelta(weeks=(current_period - 2))
current_period_end = season_start + datetime.timedelta(weeks=current_period)

# Print resulting calculations for debugging purposes
log_print(f"Current week: {current_week}", logging.DEBUG)
log_print(f"Current period: {current_period}", logging.DEBUG)
log_print(f"Current period start: {current_period_start}", logging.DEBUG)
log_print(f"Current period end: {current_period_end}", logging.DEBUG)

# Fetch forum data for a given ID
async def fetch_forum_threads(session, anime_id):
    # Start by accessing the episode discussion subforum keyed to anime ID
    base_url = 'https://myanimelist.net/forum/?animeid={}&topic=episode'
    full_base_url = 'https://myanimelist.net{}'
    url = base_url.format(anime_id)
    log_print(f"Fetching URL: {url}", VERBOSE)
    
    for attempt in range(max(1, retry_limit)):
        try:
            async with session.get(url) as response:
                if response.status != 200:
                    raise Exception(f"Failed to fetch {url} with status code: {response.status}")
                
                log_print(f"Successfully fetched data for anime_id: {anime_id}", logging.DEBUG)
        
                text = await response.text()
                soup = BeautifulSoup(markup=text, features='html.parser')
                thread_data = []
                # Parse for discussion threads and note the episode number
                pattern = re.compile(r'Episode (\d+) Discussion', re.IGNORECASE)
                threads = soup.find_all('td', class_='forum_boardrow1')
                
                log_print(f"Number of threads found: {len(threads)}", logging.DEBUG)

                if not threads:
                    log_print(f"No threads found for anime_id: {anime_id}", VERBOSE)
                    return []
                
                for thread in threads:
                    if thread.get('align') == 'right':
                        continue
                        
                    a_tag = thread.find('a', string=pattern)
                    if a_tag:
                        episode_number = pattern.search(a_tag.text).group(1)
                        thread_url = full_base_url.format(a_tag['href'])
                        reply_count_tag = thread.find_next('td')
                        
                        if reply_count_tag:
                            reply_count = int(reply_count_tag.text.strip())
                        else:
                            reply_count = 0
                        
                        thread_data.append({
                            'anime_id': anime_id,
                            'episode_number': int(episode_number),
                            'thread_url': thread_url,
                            'reply_count': reply_count
                        })
                    else:
                        log_print(f"No matching discussion thread found in this row: {thread}", VERBOSE)
                log_print(f"Number of thread_data entries: {len(thread_data)}", logging.DEBUG)
                return thread_data
        except Exception as e:
            log_print(f"Error in fetch_forum_threads() for anime ID {anime_id}: {str(e)}", logging.ERROR)
            if attempt == max(0, retry_limit - 1):
                log_print(f"Failed to fetch forum threads for anime ID {anime_id} after {retry_limit} attempts: {str(e)}", logging.ERROR)
                return []
            await asyncio.sleep(sleep_time)
    return []

# Function to fetch pages
async def fetch_page(session, url, headers):
    for attempt in range(max(1, retry_limit)):
        try:
            async with session.get(url, headers=headers) as response:
                log_print(f"Attempt {attempt + 1}: Status code {response.status} for URL: {url}", VERBOSE)
                if response.status == 403:
                    log_print(f"Access denied for {url} with status code: {response.status}. Retrying...", VERBOSE)
                    await asyncio.sleep(1)
                    continue
                elif response.status != 200:
                    log_print(f"Failed to fetch {url} with status code: {response.status}", logging.ERROR)
                    response_text = await response.text()
                    log_print(f"Response text: {response_text}", logging.DEBUG)
                    raise Exception(f"Failed to fetch {url} with status code: {response.status}")
                
                response_text = await response.text()
                log_print(f"Response text: {response_text}", logging.DEBUG)
                try:
                    json_response = await response.json()
                    return json_response
                except json.JSONDecodeError as json_error:
                    log_print(f"Failed to decode JSON from response: {str(json_error)}", logging.ERROR)
                    log_print(f"Raw response: {response_text}", logging.DEBUG)
                    raise
        except Exception as e:
            log_print(f"Error fetching {url}: {str(e)}", logging.ERROR)
            if attempt == max(0, retry_limit - 1):
                log_print(f"Failed to fetch page {url} after {retry_limit} attempts: {str(e)}", logging.ERROR)
                return None
            await asyncio.sleep(sleep_time)
    return None

# Fetch text contents of a thread via API in batches
async def fetch_thread_details(session, topicid, batch_size=100):
    base_url = f'https://api.myanimelist.net/v2/forum/topic/{topicid}'
    headers = {'X-MAL-CLIENT-ID': client_id}
    log_print(f"Using client ID: {'*' * (len(client_id) - 4) + client_id[-4:]}", logging.DEBUG)
    thread_details = []
    offset = 0

    while True:
        url = f"{base_url}?offset={offset}&limit={batch_size}"
        log_print(f"Fetching thread details from {url}", VERBOSE)
        response = await fetch_page(session, url, headers)
        if response is None:
            log_print(f"No response received for topicid {topicid}", logging.ERROR)
            break

        log_print(f"Response for topicid {topicid}: {response}", logging.DEBUG)

        posts = response.get('data', {}).get('posts', [])
        if not posts:
            log_print(f"No posts found in response for topicid {topicid}", logging.WARNING)        

        # Collect poster name and timestamp
        for post in posts:
            created_at = post.get('created_at')
            created_by = post.get('created_by', {}).get('name')
            if created_at and created_by:
                timestamp = datetime.datetime.fromisoformat(created_at.replace('Z', '+00:00')).astimezone(datetime.timezone(datetime.timedelta(hours=2)))
                thread_details.append({'username': created_by, 'timestamp': timestamp.isoformat()})
            else:
                log_print(f"Incomplete post data for topicid {topicid}: {post}", logging.WARNING)

        # Go through each page of a thread
        paging = response.get('paging', {})
        next_page = paging.get('next')
        if next_page:
            offset += batch_size
        else:
            break
    
    log_print(f"Thread details for topicid {topicid}: {thread_details}", logging.DEBUG)
    return thread_details

# Function to scrape the last post which is not exposed through API
async def fetch_last_post(session, thread_url):
    last_post_url = f"{thread_url}&goto=lastpost"
    log_print(f"Fetching last post from {last_post_url}", VERBOSE)
    
    for attempt in range(retry_limit):
        try:
            async with session.get(last_post_url) as response:
                if response.status != 200:
                    log_print(f"Failed to fetch {last_post_url} with status code: {response.status}", logging.INFO)
                    return None
                
                text = await response.text()
                soup = BeautifulSoup(text, 'html.parser')
                
                last_post_div = soup.find('div', class_='forum-topic-message message individual')
                if not last_post_div:
                    log_print(f"No last post found for {last_post_url}", VERBOSE)
                    return None
                
                username = last_post_div.get('data-user')
                timestamp_div = last_post_div.find_next('div', class_='date')
                timestamp_unix = int(timestamp_div.get('data-time'))
                timestamp = datetime.datetime.fromtimestamp(timestamp_unix, datetime.timezone.utc).astimezone(datetime.timezone(datetime.timedelta(hours=2)))
                
                last_post_details = {'username': username, 'timestamp': timestamp.isoformat()}
                log_print(f"Last post details for {last_post_url}: {last_post_details}", logging.DEBUG)
                return last_post_details
        except Exception as e:
            if attempt == retry_limit - 1:
                log_print(f"Failed to fetch last post from {last_post_url} after {retry_limit} attempts: {str(e)}", logging.ERROR)
                return None
            await asyncio.sleep(sleep_time)

async def fetch_forum_data(anime_ids):
    async with get_session() as session:
        # Fetch all forum threads
        thread_results = await asyncio.gather(*[fetch_forum_threads(session, anime_id) for anime_id in anime_ids])
        all_thread_data = [thread for result in thread_results if result is not None for thread in result]
        
        if not all_thread_data:
            log_print("No thread data found for any anime.", logging.INFO)
            return []
        else:
            log_print(f"Total threads found: {len(all_thread_data)}", logging.DEBUG)
            for thread in all_thread_data:
                log_print(f"Thread data: {thread}", logging.DEBUG)
        
        # Fetch all thread details and last posts concurrently
        detail_tasks = [fetch_thread_details(session, thread['thread_url'].split('=')[-1]) for thread in all_thread_data]
        last_post_tasks = [fetch_last_post(session, thread['thread_url']) for thread in all_thread_data]
        
        all_results = await asyncio.gather(*detail_tasks, *last_post_tasks)
        
        detail_results = all_results[:len(detail_tasks)]
        last_post_results = all_results[len(detail_tasks):]

        for thread, details, last_post in zip(all_thread_data, detail_results, last_post_results):
            thread['details'] = details
            if last_post:
                thread['details'].append(last_post)

        # Collect everything we've gathered into a dictionary
        anime_thread_data = []
        for anime_id in anime_ids:
            anime_data = {
                'anime_id': anime_id,
                'threads': []
            }
            for thread in all_thread_data:
                if thread['anime_id'] == anime_id:
                    unique_posters = set(detail['username'] for detail in thread['details'])
                    thread['unique_posters'] = unique_posters
                    anime_data['threads'].append(thread)
            anime_thread_data.append(anime_data)

        return anime_thread_data

# Function to create and populate the posts sheet if we're fetching forum data
def create_posts_sheet(wb, data, titles):
    ws = wb.create_sheet(title="posts")
    post_counts = {}

    # Add the first heading
    ws.append(["Total number of unique posters in episode discussion threads"])
    ws.cell(row=1, column=1).font = Font(size=18)

    # Put the post summary
    all_episode_numbers = sorted(set(
        thread['episode_number']
        for anime_data in data
        for thread in anime_data['threads']
    ))
    header = ["Title"] + [f"EP{ep}" for ep in all_episode_numbers]
    ws.append(header)

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for anime_data in data:
        anime_id = anime_data['anime_id']
        title = titles.get(anime_id, f"Unknown (ID: {anime_id})")
        log_print(f"\nProcessing anime {title} (ID: {anime_id})", logging.DEBUG)
        row = [title]
        episode_dict = {thread['episode_number']: len(thread['unique_posters']) for thread in anime_data['threads']}
        for ep in all_episode_numbers:
            row.append(episode_dict.get(ep, 0))
        ws.append(row)

        # Calculate unique posters
        total_posts = 0
        for thread in anime_data['threads']:
            log_print(f"Processing thread for EP{thread['episode_number']}", logging.DEBUG)
            thread_posters = set()
            for detail in thread['details']:
                timestamp = datetime.datetime.fromisoformat(detail['timestamp'])
                log_print(f"Post timestamp: {timestamp}", logging.DEBUG)
                # For week 2, count all posts made before its end
                if current_period == 2:
                    if timestamp < current_period_end:
                        thread_posters.add(detail['username'])
                        log_print(f"Added poster {detail['username']} to thread set", logging.DEBUG)
                else:
                    if current_period_start <= timestamp < current_period_end:
                        thread_posters.add(detail['username'])
                        log_print(f"Added poster {detail['username']} to thread set", logging.DEBUG)
            
            # Sum up unique posts for each thread
            total_posts += len(thread_posters)
            log_print(f"Unique posters in thread for EP{thread['episode_number']}: {len(thread_posters)}", logging.DEBUG)
        
        # Sum up unique posts across all threads keyed to the anime ID
        post_counts[anime_id] = total_posts
        log_print(f"Total posts for {title}: {post_counts[anime_id]}", logging.DEBUG)

    # Add second heading before the thread breakdown
    ws.append([])
    ws.append(["Unique poster breakdown per thread (name, number of posts, date of first post)"])
    ws.cell(row=ws.max_row, column=1).font = Font(size=18)
    
    breakdown_start_row = ws.max_row + 1  # Remember where the breakdown starts
    
    for anime_data in data:
        anime_id = anime_data['anime_id']
        title = titles[anime_id]
        
        # Define the structure for per-episode breakdown
        for i, thread in enumerate(anime_data['threads']):
            if i > 0 or anime_data != data[0]:
                ws.append([])
            ws.append([title, f"EP{thread['episode_number']}", thread['thread_url']])
            row_index = ws.max_row
            for cell in ws[row_index]:
                cell.font = Font(bold=True)
            ws[row_index][1].alignment = Alignment(horizontal='center')
            ws[row_index][2].hyperlink = thread['thread_url']
            ws[row_index][2].style = 'Hyperlink'
            
            user_posts = defaultdict(lambda: {'count': 0, 'first_post': None})
            for detail in thread['details']:
                username = detail['username']
                timestamp = datetime.datetime.fromisoformat(detail['timestamp'])
                user_posts[username]['count'] += 1
                if not user_posts[username]['first_post']:
                    user_posts[username]['first_post'] = timestamp.strftime('%Y-%m-%d %H:%M')
            
            # Sort the user list by total post counts, then first post date
            sorted_user_posts = sorted(user_posts.items(), key=lambda x: x[1]['count'], reverse=True)
            for username, info in sorted_user_posts:
                ws.append([username, info['count'], info['first_post']])
        
    ws.column_dimensions['A'].width = 2 * ws.column_dimensions['A'].width

    # Apply number format only to the user post counts
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = numbers.FORMAT_NUMBER

    # Apply gradient conditional formatting to user post counts if they exist
    if ws.max_row > breakdown_start_row:
        last_data_row = ws.max_row
        color_rule_posts = ColorScaleRule(
            start_type='num', start_value=1, start_color='209020',  # Forest green
            mid_type='num', mid_value=5, mid_color='F0F050',  # Starship yellow
            end_type='num', end_value=15, end_color='C00000'  # Dark red
        )
        # Apply conditional formatting after the second heading
        ws.conditional_formatting.add(f'B{breakdown_start_row}:B{last_data_row}', color_rule_posts)
    else:
        log_print("No data to apply conditional formatting in the posts sheet.", logging.WARNING)
    return post_counts

# Add text summary displaying internal timestamps below data rows
main_sheet.append([])
main_sheet.append([
    f"Current season: {current_season} {now.year}, data retrieved on {now.strftime('%Y-%m-%d %H:%M')} (FAL timezone)."
])
main_sheet.append([
    f"Retrieval week: {current_week}, post counting period: week {current_period}."
])

# Fetch and integrate forum data into the workbook if enabled
if enable_posts:
    forum_data = asyncio.run(fetch_forum_data(anime_ids))
    anime_titles = {}
    for anime_id in anime_ids:  # Use the original list of anime IDs
        data = datas.get(anime_id)
        if isinstance(data, dict) and 'title' in data:
            anime_titles[anime_id] = data['title']
        else:
            log_print(f"Warning: Missing or invalid data for anime ID {anime_id}", logging.WARNING)
            anime_titles[anime_id] = f"Unknown (ID: {anime_id})"
    post_counts = create_posts_sheet(workbook, forum_data, anime_titles)

    log_print("\nPopulating Posts column", VERBOSE)
    main_sheet = workbook['main']
    for row in main_sheet.iter_rows(min_row=2, max_row=len(datas) + 1, min_col=4, max_col=4):
        cell = row[0]
        anime_id = main_sheet.cell(row=cell.row, column=13).value
        post_count = post_counts.get(anime_id, 0)
        log_print(f"Anime ID: {anime_id}, Post count: {post_count}", logging.DEBUG)
        cell.value = post_counts.get(anime_id, 0)    

# Timestamp output filename to identify data and avoid accidental overwrites
now = datetime.datetime.now()
timestamp = now.strftime("%Y-%m-%d-%H-%M")
filename = f"FAL_data_{season_code}_{timestamp}.xlsx"
workbook.save(filename)
log_print(f"Written to file: {filename}", logging.INFO)

# Write the final log items and provide the error summary, if any
log_print(f"Finished processing with {error_count} errors", logging.INFO)
if error_count > 0:
    log_print(f"Failed to fetch data for the following anime IDs: {failed_anime_ids}", VERBOSE)