import datetime
import requests
from openpyxl import Workbook

# Insert the client ID you received from MAL in the quotes.
CLIENT_ID = ''

# Populate this list with IDs of the anime that need tracking. The ID can be found in the page URL of the anime in question, i.e. for Trigun (https://myanimelist.net/anime/6/Trigun) the ID is "6". The following list is a placeholder only relevant for spring 2023.
ids = [45486, 50220, 50307, 50416, 50796, 51219, 51614, 51632, 51705, 51706, 51817, 51958, 52034, 52092, 52211, 52308, 52578, 52608, 52657, 52830, 52955, 53126, 53199, 53393, 53613]
datas = []

for i in ids:
    url = 'https://api.myanimelist.net/v2/anime/' + str(i) + '?fields=mean,num_favorites,statistics'
    response = requests.get(url, headers = {
        'X-MAL-CLIENT-ID': CLIENT_ID
        })

    response.raise_for_status()
    anime = response.json()

    print(anime)
    datas.append(anime)
response.close()

# Setting up the Excel file for the output. Feel free to rename column headers.
headers = ['Title', 'Score', 'Favorites', 'Watching', 'W+C', 'Dropped', 'PTW']
workbook = Workbook()
sheet = workbook.active

for i, header in enumerate(headers):
    sheet.cell(1, i+1).value = header

for j, data in enumerate(datas):
    sheet.cell(j+2, 1).value = str(data['title'])
    sheet.cell(j+2, 2).value = data.get("mean") or '-'
    sheet.cell(j+2, 3).value = data['num_favorites']
    stats = data['statistics']['status']
    sheet.cell(j+2, 4).value = stats['watching']
    # Calculate the sum of 'watching' and 'completed' for the 'W+C' column.
    sheet.cell(j+2, 5).value = int(stats['watching']) + int(stats['completed'])
    sheet.cell(j+2, 6).value = stats['dropped']
    sheet.cell(j+2, 7).value = stats['plan_to_watch']

# Fix the value type for cells which Excel insists on treating as text, messing up formulas.
for column_num in range(4, 8):
    for row_num in range(2, sheet.max_row+1):
        cell = sheet.cell(row_num, column_num)
        if cell.value is not None and cell.value != '-':
            cell.value = int(cell.value) 

# Write a timestamp into the output filename so the files are easily identified and aren't overwritten by repeated usage of the script.
now = datetime.datetime.now()
timestamp = now.strftime("%Y-%m-%d-%H-%M")
filename = f"FAL_data_{timestamp}.xlsx"
workbook.save(filename)

# Made by eplipswich and moozooh.