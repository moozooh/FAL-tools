import datetime
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import GradientFill, PatternFill, Alignment, Font, colors, numbers
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# Insert the client ID you received from MAL in the quotes.
CLIENT_ID = ''

# Populate this list with IDs of the anime that need tracking. The ID can be found in the page URL of the anime in question, i.e. for Trigun (https://myanimelist.net/anime/6/Trigun) the ID is "6". The following list is a placeholder only relevant for spring 2023.
ids = [45486, 50220, 50307, 50416, 50796, 51219, 51614, 51632, 51705, 51706, 51817, 51958, 52034, 52092, 52211, 52308, 52578, 52608, 52657, 52830, 52955, 53126, 53199, 53393, 53613]
datas = []

# Grab data from the API.
for i in ids:
    url = 'https://api.myanimelist.net/v2/anime/' + str(i) + '?fields=mean,num_favorites,statistics'
    response = requests.get(url, headers = {
        'X-MAL-CLIENT-ID': CLIENT_ID
        })

    response.raise_for_status()
    anime = response.json()

    print(anime)
    datas.append(anime)
response.close()

# Set up the output Excel file. Feel free to rename column headers.
headers = ['Title', 'Score', 'Favorites', 'Watching', 'W+C', 'Dropped', 'Drop Rate', 'PTW', 'PTW Ratio']
workbook = Workbook()
sheet = workbook.active

# Align the headers and make them bold. Freeze headers and first column.
for i, header in enumerate(headers):
    sheet.cell(1, i+1).value = header
    sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
    sheet.cell(1, i+1).font = Font(bold=True)
    sheet.freeze_panes = 'B2'

# Populate the data rows.
for j, data in enumerate(datas):
    sheet.cell(j+2, 1).value = str(data['title'])
    sheet.cell(j+2, 2).value = data.get("mean") or ' '
    sheet.cell(j+2, 2).number_format = numbers.FORMAT_NUMBER_00
    sheet.cell(j+2, 3).value = int(data['num_favorites'])
    stats = data['statistics']['status']
    sheet.cell(j+2, 4).value = int(stats['watching'])
    # Calculate the sum of 'watching' and 'completed' for the 'W+C' column.
    watch_comp = int(stats['watching']) + int(stats['completed'])
    sheet.cell(j+2, 5).value = watch_comp
    sheet.cell(j+2, 6).value = int(stats['dropped'])
    # Calculate drop rate and display it as a percentage.
    sheet.cell(j+2, 7).value = int(stats['dropped']) / int(stats['watching'])
    sheet.cell(j+2, 7).number_format = numbers.FORMAT_PERCENTAGE_00
    ptw = int(stats['plan_to_watch'])
    sheet.cell(j+2, 8).value = ptw
    # Calculate ratio of PTW to active audience and display it as a percentage. Disregard W+C values under 400 to avoid false positives.
    if watch_comp >= 400:
        active_ratio = watch_comp / ptw
    else:
        active_ratio = 0.00
    sheet.cell(j+2, 9).value = active_ratio
    sheet.cell(j+2, 9).number_format = numbers.FORMAT_PERCENTAGE_00

# Apply conditional green-white-red gradient formatting to percentage columns.
# For column G, values ≤0.3% are green, ≥10% are red, and 3.5% is the midpoint.
green_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
red_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
color_rule_g = ColorScaleRule(start_type='num', start_value=0.003, start_color='548235', mid_type='num', mid_value=0.035, mid_color='FFFFFF', end_type='num', end_value=0.1, end_color='C00000')
range_g = 'G2:G101'
sheet.conditional_formatting.add(range_g, color_rule_g)
sheet.conditional_formatting.add(range_g, CellIsRule(operator='greaterThanOrEqual', formula=['0.1'], fill=red_fill))
# Skip empty cells and values equal to 0%.
for row in sheet.iter_rows(min_row=2, max_row=101, min_col=7, max_col=7):
    for cell in row:
        if cell.value is not None and cell.value > 0:
            sheet.conditional_formatting.add(cell.coordinate, CellIsRule(operator='lessThanOrEqual', formula=['0.003'], fill=green_fill))

# For column I, the lowest is green, the highest red, and 100% is the midpoint.
color_rule_i = ColorScaleRule(start_type='min', start_color='548235', end_type='max', end_color='C00000', mid_type='num', mid_value=1.0, mid_color='FFFFFF')
range_i = 'I2:I101'
for row in sheet.iter_rows(min_row=2, max_row=101, min_col=9, max_col=9):
    for cell in row:
        if cell.value is not None and cell.value > 0:
            sheet.conditional_formatting.add(range_i, color_rule_i)

# Add comments explaining columns and their formatting.
cell_e1 = sheet['E1']
comment_e1 = openpyxl.comments.Comment("Sum of watching and completed users.", " ")
comment_e1.height = None
comment_e1.width = None
cell_e1.comment = comment_e1
cell_g1 = sheet['G1']
comment_g1 = openpyxl.comments.Comment("Number of dropped users taken as a percentage of watching + completed users.\nLower values are better. Peak positive values are 0.30% and below. Peak negative values are 10.00% and above. Midpoint is at 3.50%.", " ")
comment_g1.height = None
comment_g1.width = None
cell_g1.comment = comment_g1
cell_i1 = sheet['I1']
comment_i1 = openpyxl.comments.Comment("Ratio of watching + completed users (active audience) to PTW users (potential audience).\nLower values are better. Higher values may indicate upcoming slowdown in growth as series run out of potential audience to convert into active.\nMidpoint is at 100%.", " ")
comment_i1.height = None
comment_i1.width = None
cell_i1.comment = comment_i1

# Timestamp output filenames to identify data easily and avoid overwriting it by repeated usage of the script.
now = datetime.datetime.now()
timestamp = now.strftime("%Y-%m-%d-%H-%M")
filename = f"FAL_data_{timestamp}.xlsx"
workbook.save(filename)